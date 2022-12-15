import openpyxl
import os


# Read the input file name from the user
input_file = input('Enter the name of the input file: ')
input_file += '.xlsx'

# Open the Excel file
try:
  wb = openpyxl.load_workbook(input_file)
except FileNotFoundError:
  print('The input file could not be found.')
  exit()

# Get the first sheet in the workbook
sheet = wb["Sheet1"]

# Read the username from the user
username = input('Enter the username: ')

# Concatenate the @conectys.com string to the username
username += '@conectys.com'

# Create a new workbook for the output
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active

# Set the column headers for the output sheet
output_sheet['A1'] = 'tcs_link'
output_sheet['B1'] = '1_verifier'
output_sheet['C1'] = 'classification'
output_sheet['D1'] = '1_item_clickbait_label'
output_sheet['E1'] = 'QA_verifier'
output_sheet['F1'] = 'QA_classification'
output_sheet['G1'] = 'QA_item_clickbait_label'

# Loop through the rows in the input sheet
for i in range(2, sheet.max_row + 1):
  # Check if the username matches the value in the 2nd column
  if sheet.cell(row=i, column=2).value == username:
    # If it matches, add the values from the 11th, 2nd, 3rd, and 4th columns to the output sheet
    output_sheet.cell(row=i, column=1).value = sheet.cell(row=i, column=11).value
    output_sheet.cell(row=i, column=2).value = sheet.cell(row=i, column=2).value
    output_sheet.cell(row=i, column=3).value = sheet.cell(row=i, column=3).value
    output_sheet.cell(row=i, column=4).value = sheet.cell(row=i, column=4).value
    output_sheet.cell(row=i, column=5).value = sheet.cell(row=i, column=8).value
    output_sheet.cell(row=i, column=6).value = sheet.cell(row=i, column=9).value
    output_sheet.cell(row=i, column=7).value = sheet.cell(row=i, column=10).value




# Save the output workbook
output_wb.save('output.xlsx')
print('1_verifier column data has extracted to', os.path.abspath('output.xlsx'))

# Loop through the rows in the input sheet again
for i in range(2, sheet.max_row + 1):
  # Check if the username matches the value in the 5th column
  if sheet.cell(row=i, column=5).value == username:
    # If it matches, add the values from the 11th, 5th, 6th, and 7th columns to the output sheet
    output_sheet.cell(row=i, column=1).value = sheet.cell(row=i, column=11).value
    output_sheet.cell(row=i, column=2).value = sheet.cell(row=i, column=5).value
    output_sheet.cell(row=i, column=3).value = sheet.cell(row=i, column=6).value
    output_sheet.cell(row=i, column=4).value = sheet.cell(row=i, column=7).value
    output_sheet.cell(row=i, column=5).value = sheet.cell(row=i, column=8).value
    output_sheet.cell(row=i, column=6).value = sheet.cell(row=i, column=9).value
    output_sheet.cell(row=i, column=7).value = sheet.cell(row=i, column=10).value

# Save the output workbook
output_wb.save('output.xlsx')
print('2_verifier column data has extracted to', os.path.abspath('output.xlsx'))

